"""
Создание СОВСЕМ НОВЫХ карточек на WB (которых раньше не было в продаже) —
по тому же принципу "по образцу", что и ozon_new_products.py: вместо ручного
выбора раздела (subject) и заполнения десятков характеристик, вы указываете
артикул уже существующего похожего товара — скрипт берёт у него раздел,
бренд, характеристики и размеры.

ВАЖНО, чем это отличается от обычного редактирования (wb_catalog_editor.py):
- Метод создания (/content/v2/cards/upload) — ДРУГОЙ, не тот, что для правки
  существующих карточек. Он асинхронный и не сразу возвращает nmID нового
  товара.
- Штрихкод для нового размера генерируется автоматически (у WB есть для
  этого отдельный метод) — свой штрихкод указывать не нужно.
- ФОТО И ВИДЕО сюда не входят — их нельзя отправить без nmID, а WB не
  отдаёт nmID сразу при создании. Порядок действий:
    1. build-wb-new-template — заполнить таблицу.
    2. push-wb-new-cards-dryrun — посмотреть, что будет отправлено.
    3. push-wb-new-cards — реально создать карточки (пока без фото/видео).
    4. Подождать 2-5 минут, затем fetch-wb ещё раз — новые товары появятся
       в data/wb_cards.json уже со своим nmID.
    5. Дальше — как с обычными товарами: build-wb-catalog, attach-wb-photos,
       push-wb-cards (там же можно сразу дозаполнить видео).

НЕ ПРОВЕРЕНО на реальных данных: точные названия полей раздела (subjectID)
и структуры "dimensions"/"sizes" в ответе WB — проверьте на первом же
тестовом товаре через push-wb-new-cards-dryrun и напишите, если что-то в
выводе выглядит не так, прежде чем запускать push-wb-new-cards.
"""
import logging
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger("marketplace-agent.wb_new")

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")

COLUMNS = [
    ("vendor_code", "НОВЫЙ артикул WB (vendorCode) — придумайте сами, уникальный"),
    ("sample_vendor_code", "Образец: артикул похожего товара, уже продающегося на WB"),
    ("title", "Название (до 60 символов)"),
    ("description", "Описание"),
    ("price", "Цена, ₽"),
    ("weight_kg", "Вес, кг (пусто = как у образца)"),
    ("length_cm", "Длина, см (пусто = как у образца)"),
    ("width_cm", "Ширина, см (пусто = как у образца)"),
    ("height_cm", "Высота, см (пусто = как у образца)"),
    ("notes", "Заметки"),
]

WIDTHS = [20, 30, 45, 45, 12, 14, 14, 14, 14, 25]


def build_new_template(xlsx_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Новые товары WB"

    for col_idx, (_, header) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for col_idx, width in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(xlsx_path)


def load_new_edits(xlsx_path: str) -> Dict[str, dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    edits: Dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        vendor_code = str(row[0]).strip()
        raw = {COLUMNS[i][0]: (row[i] if i < len(row) else None) for i in range(len(COLUMNS))}
        edits[vendor_code] = {
            "sample_vendor_code": str(raw.get("sample_vendor_code") or "").strip(),
            "title": (raw.get("title") or "").strip() if isinstance(raw.get("title"), str) else raw.get("title"),
            "description": raw.get("description") or "",
            "price": raw.get("price"),
            "weight_kg": raw.get("weight_kg"),
            "length_cm": raw.get("length_cm"),
            "width_cm": raw.get("width_cm"),
            "height_cm": raw.get("height_cm"),
            "notes": raw.get("notes"),
        }
    return edits


def _num_or(value, default):
    if value in (None, ""):
        return default
    try:
        return type(default)(value) if default not in (None, "") else value
    except (TypeError, ValueError):
        return default


def resolve_subject_id(sample_card: dict) -> Optional[int]:
    """
    Раздел (subject) нужен ЧИСЛОМ для создания карточки. Если он уже есть в
    уже выгруженных данных образца — берём как есть. Если нет (в списке
    карточек WB может отдавать только текстовое название раздела) — ищем по
    названию через живой список разделов WB.
    """
    subject_id = sample_card.get("subjectID")
    if subject_id:
        return subject_id

    name = sample_card.get("subjectName")
    if not name:
        return None

    import wb_client

    try:
        subjects = wb_client.get_subjects(name=name)
    except wb_client.WbApiError as exc:
        logger.warning("Не удалось найти раздел WB по названию '%s': %s", name, exc)
        return None

    for s in subjects:
        candidate_name = s.get("subjectName") or s.get("name") or ""
        if candidate_name.strip().lower() == str(name).strip().lower():
            return s.get("subjectID") or s.get("id")
    if subjects:
        # Точного совпадения не нашлось — берём первый результат поиска, но
        # предупреждаем, чтобы проверили вручную перед push-wb-new-cards.
        logger.warning(
            "Раздел '%s' не совпал точно ни с одним результатом поиска — "
            "взят первый похожий результат, проверьте перед реальной отправкой.",
            name,
        )
        return subjects[0].get("subjectID") or subjects[0].get("id")
    return None


def build_new_card_groups(wb_data: dict, edits: Dict[str, dict], barcodes: List[str]) -> List[dict]:
    """
    Строит группы для wb_client.create_cards (один вызов на один subjectID):
    [{"subject_id": int, "vendor_code": str, "variants": [...]}, ...] —
    группировка НЕ по subjectID здесь, а по одной карточке за раз (проще и
    безопаснее следить за ошибками по каждому товару отдельно, чем одним
    большим пакетом на весь subjectID). barcodes — заранее сгенерированный
    список уникальных штрихкодов, по одному на каждый новый товар (порядок
    соответствует порядку строк edits).
    """
    from wb_catalog_editor import _cards_by_vendor

    cards = _cards_by_vendor(wb_data)
    groups: List[dict] = []
    barcode_iter = iter(barcodes)

    for vendor_code, edit in edits.items():
        sample_code = edit.get("sample_vendor_code")
        sample = cards.get(sample_code)
        if not sample:
            logger.warning(
                "%s: образец '%s' не найден в data/wb_cards.json (сначала fetch-wb, "
                "проверьте артикул образца без опечаток) — товар пропущен.",
                vendor_code,
                sample_code,
            )
            continue

        subject_id = resolve_subject_id(sample)
        if not subject_id:
            logger.warning(
                "%s: не удалось определить раздел (subject) по образцу '%s' — товар пропущен.",
                vendor_code,
                sample_code,
            )
            continue

        try:
            barcode = next(barcode_iter)
        except StopIteration:
            logger.warning("%s: не хватило сгенерированных штрихкодов — товар пропущен.", vendor_code)
            continue

        sample_sizes = sample.get("sizes") or [{}]
        sample_size = sample_sizes[0] if sample_sizes else {}
        price = edit.get("price")
        price = price if price not in (None, "") else sample_size.get("price", 0)
        # Берём у образца только техразмер (для настоящих деталей обычно
        # заглушка "0"/пусто) — служебные поля вроде chrtID (внутренний ID
        # WB для УЖЕ СУЩЕСТВУЮЩЕГО размера образца) переиспользовать нельзя,
        # для нового товара их не отправляем вообще.
        new_size = {
            "techSize": sample_size.get("techSize", "0"),
            "wbSize": sample_size.get("wbSize", ""),
            "price": price,
            "skus": [barcode],
        }

        dimensions_raw = dict(sample.get("dimensions") or {})
        dimensions = {
            "length": dimensions_raw.get("length", 0),
            "width": dimensions_raw.get("width", 0),
            "height": dimensions_raw.get("height", 0),
            "weightBrutto": dimensions_raw.get("weightBrutto", 0),
        }
        if edit.get("length_cm") not in (None, ""):
            dimensions["length"] = edit.get("length_cm")
        if edit.get("width_cm") not in (None, ""):
            dimensions["width"] = edit.get("width_cm")
        if edit.get("height_cm") not in (None, ""):
            dimensions["height"] = edit.get("height_cm")
        if edit.get("weight_kg") not in (None, ""):
            dimensions["weightBrutto"] = edit.get("weight_kg")

        variant = {
            "vendorCode": vendor_code,
            "title": (edit.get("title") or "")[:60],
            "description": edit.get("description") or "",
            "brand": sample.get("brand", ""),
            "dimensions": dimensions,
            "characteristics": sample.get("characteristics") or [],
            "sizes": [new_size],
        }
        groups.append({"subject_id": subject_id, "vendor_code": vendor_code, "variant": variant})

    return groups
