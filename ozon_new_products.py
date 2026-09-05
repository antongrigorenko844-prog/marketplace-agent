"""
Создание СОВСЕМ НОВЫХ карточек на Ozon (которых раньше не было в продаже) —
в отличие от catalog_editor.py, который только редактирует уже существующие
товары.

Как это работает (принцип "по образцу"):
У новой запчасти нет своей категории/характеристик/размеров — их пришлось
бы выбирать вручную из огромного дерева категорий Ozon. Вместо этого вы
указываете АРТИКУЛ УЖЕ СУЩЕСТВУЮЩЕГО похожего товара ("образец") — скрипт
берёт у него категорию, тип, характеристики, вес и размеры и использует их
для нового товара, подставляя только то, что вы заполнили сами (артикул,
название, описание, цена, фото, видео). Вес/размеры можно переопределить
отдельно, если у нового товара они другие.

ВАЖНО: это создаёт РЕАЛЬНЫЙ новый товар в Ozon (тот же метод /v3/product/import,
что и для правок — Ozon сам решает, создать новый или обновить существующий,
по значению offer_id). Если офер_id уже существует — товар будет ОБНОВЛЁН, а
не создан заново, поэтому для новых товаров offer_id должен быть новым,
которого раньше не было.

Порядок команд:
1. build-ozon-new-template — пустая таблица data/ozon_new_products.xlsx.
2. Заполнить строки (артикул, образец, название, описание, цена...).
3. attach-ozon-new-photos — подставляет фото из папки photos/ так же, как
   для обычных товаров, но по НОВОМУ артикулу.
4. push-ozon-new-cards-dryrun — посмотреть, что будет отправлено.
5. push-ozon-new-cards — реально создать товары.

После успешного создания — выполните fetch-ozon ещё раз: новые товары
попадут в общий data/ozon_products.json и дальше редактируются как обычные,
через build-ozon-catalog.
"""
import logging
import os
from typing import Dict, List, Optional
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import catalog_editor

logger = logging.getLogger("marketplace-agent.ozon_new")

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name=FONT_NAME)

COLUMNS = [
    ("offer_id", "НОВЫЙ артикул (offer_id) — придумайте сами, уникальный"),
    ("sample_offer_id", "Образец: артикул похожего товара, уже продающегося на Ozon"),
    ("name", "Название"),
    ("description", "Описание"),
    ("price", "Цена, ₽"),
    ("old_price", "Цена до скидки, ₽ (необязательно)"),
    ("barcode", "Штрихкод (необязательно — можно оставить пустым)"),
    ("weight_g", "Вес, г (пусто = как у образца)"),
    ("length_mm", "Длина, мм (пусто = как у образца)"),
    ("width_mm", "Ширина, мм (пусто = как у образца)"),
    ("height_mm", "Высота, мм (пусто = как у образца)"),
    ("images", "Фото: ссылки через | (заполняется автоматически attach-ozon-new-photos)"),
    ("video_url", "Видео: ссылка на .mp4/.mov, необязательно, ЭКСПЕРИМЕНТАЛЬНО"),
    ("notes", "Заметки"),
]

WIDTHS = [20, 30, 45, 45, 12, 18, 20, 14, 14, 14, 14, 55, 45, 25]


def build_new_template(xlsx_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Новые товары Ozon"

    for col_idx, (_, header) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for col_idx, width in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(xlsx_path)


def load_new_edits(xlsx_path: str) -> Dict[str, dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    edits: Dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        offer_id = str(row[0]).strip()
        raw = {COLUMNS[i][0]: (row[i] if i < len(row) else None) for i in range(len(COLUMNS))}
        images_raw = raw.get("images") or ""
        images = [u.strip() for u in str(images_raw).split("|") if u.strip()]
        edits[offer_id] = {
            "sample_offer_id": str(raw.get("sample_offer_id") or "").strip(),
            "name": (raw.get("name") or "").strip() if isinstance(raw.get("name"), str) else raw.get("name"),
            "description": raw.get("description") or "",
            "price": raw.get("price"),
            "old_price": raw.get("old_price"),
            "barcode": str(raw.get("barcode") or "").strip(),
            "weight_g": raw.get("weight_g"),
            "length_mm": raw.get("length_mm"),
            "width_mm": raw.get("width_mm"),
            "height_mm": raw.get("height_mm"),
            "images": images,
            "video_url": (raw.get("video_url") or "").strip() if isinstance(raw.get("video_url"), str) else "",
            "notes": raw.get("notes"),
        }
    return edits


def attach_new_photos(xlsx_path: str, photos_dir: str, raw_base_url: str) -> Dict[str, List[str]]:
    """То же самое, что catalog_editor.attach_local_photos, но по колонке offer_id этого шаблона."""
    wb = load_workbook(xlsx_path)
    ws = wb.active

    offer_id_col_idx = 1
    images_col_idx = next(i for i, (key, _) in enumerate(COLUMNS, start=1) if key == "images")

    files = [f for f in os.listdir(photos_dir) if os.path.splitext(f)[1].lower() in catalog_editor.IMAGE_EXTS]

    matched: Dict[str, List[str]] = {}
    for row in ws.iter_rows(min_row=2):
        offer_id_cell = row[offer_id_col_idx - 1]
        if not offer_id_cell.value:
            continue
        offer_id = str(offer_id_cell.value).strip()

        own_files = [
            f
            for f in files
            if f.startswith(offer_id + "_") or f == offer_id + os.path.splitext(f)[1]
        ]
        if not own_files:
            continue
        own_files.sort(key=lambda f: catalog_editor._photo_index(f, offer_id))

        urls = [f"{raw_base_url.rstrip('/')}/{quote(f)}" for f in own_files]
        row[images_col_idx - 1].value = "|".join(urls)
        matched[offer_id] = urls

    wb.save(xlsx_path)
    return matched


def _num_or(value, default):
    if value in (None, ""):
        return default
    try:
        return float(value) if isinstance(default, float) else int(value)
    except (TypeError, ValueError):
        return default


def build_new_import_items(existing_data: dict, edits: Dict[str, dict]) -> List[dict]:
    """
    Строит items для /v3/product/import для НОВЫХ товаров, клонируя
    категорию/тип/характеристики/размеры у "образца" (уже существующего
    товара из data/ozon_products.json).
    """
    names_by_offer = {n.get("offer_id"): n for n in existing_data.get("names", []) if n.get("offer_id")}

    items: List[dict] = []
    for new_offer_id, edit in edits.items():
        sample_id = edit.get("sample_offer_id")
        sample = names_by_offer.get(sample_id)
        if not sample:
            logger.warning(
                "%s: образец '%s' не найден в data/ozon_products.json (сначала fetch-ozon, "
                "и проверьте, что артикул образца указан без опечаток) — товар пропущен.",
                new_offer_id,
                sample_id,
            )
            continue

        new_name = (edit.get("name") or "").strip()
        new_description = edit.get("description") or ""

        attributes = [dict(a) for a in (sample.get("attributes") or [])]
        attributes = catalog_editor._with_overridden_attr(attributes, catalog_editor.ATTR_ID_NAME_MIRROR, new_name)
        attributes = catalog_editor._with_overridden_attr(attributes, catalog_editor.ATTR_ID_DESCRIPTION, new_description)

        video_url = (edit.get("video_url") or "").strip()
        if video_url:
            video_attr_id = catalog_editor._get_video_attr_id(
                sample.get("description_category_id", 0), sample.get("type_id", 0)
            )
            if video_attr_id:
                attributes = catalog_editor._with_overridden_attr(attributes, video_attr_id, video_url)
            else:
                logger.warning(
                    "%s: в характеристиках категории образца не нашлось атрибута 'видео' — "
                    "видео НЕ добавлено, остальное создастся как обычно.",
                    new_offer_id,
                )

        images = edit.get("images") or []
        primary_image = images[0] if images else ""
        rest_images = images[1:] if len(images) > 1 else []

        price = edit.get("price")
        price = str(price).strip() if price not in (None, "") else "0"
        old_price = edit.get("old_price")
        old_price = str(old_price).strip() if old_price not in (None, "") else "0"

        items.append(
            {
                "offer_id": new_offer_id,
                "name": new_name,
                "description_category_id": sample.get("description_category_id", 0),
                "new_description_category_id": 0,
                "type_id": sample.get("type_id", 0),
                "barcode": edit.get("barcode") or "",
                "attributes": attributes,
                "complex_attributes": sample.get("complex_attributes") or [],
                "color_image": "",
                "images": rest_images,
                "primary_image": primary_image,
                "pdf_list": [],
                "currency_code": "RUB",
                "price": price,
                "old_price": old_price,
                "vat": "0",
                "depth": _num_or(edit.get("length_mm"), sample.get("depth", 0)),
                "width": _num_or(edit.get("width_mm"), sample.get("width", 0)),
                "height": _num_or(edit.get("height_mm"), sample.get("height", 0)),
                "dimension_unit": sample.get("dimension_unit") or "mm",
                "weight": _num_or(edit.get("weight_g"), sample.get("weight", 0)),
                "weight_unit": sample.get("weight_unit") or "g",
            }
        )
    return items
