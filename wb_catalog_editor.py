"""
Построение и разбор редактируемого Excel-каталога WB — название, описание
и ФОТО. По той же схеме, что и catalog_editor.py для Ozon, но под API WB,
у которого другая механика:

- Ключ товара у WB — vendorCode (артикул продавца), но методы обновления
  карточки и фото требуют ЧИСЛОВОЙ nmID (внутренний номер WB) — он берётся
  из уже выгруженного data/wb_cards.json по vendorCode.
- /content/v2/cards/update — правит название/описание, ПЕРЕЗАПИСЫВАЕТ
  карточку целиком (нужно прислать и то, что не меняете), но ЭТОТ метод
  не умеет трогать фото/видео вообще — фото совершенно отдельный метод.
- /content/v3/media/save — правит фото/видео, тоже полная замена (новый
  список ссылок заменяет старый целиком), но никак не связан с названием.

Из-за этого у WB два независимых full-replace метода вместо одного, как у
Ozon — соответственно, ниже две отдельные функции сборки: build_wb_update_items
(текст) и build_wb_media_updates (фото), их можно отправлять по отдельности.
"""
import os
import re
from typing import Dict, List, Optional
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name=FONT_NAME)
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

IMAGE_EXTS = {".jpg", ".jpeg", ".png"}
TITLE_MAX_LEN = 60  # ограничение WB (проверено в живой документации 02.09.2026)

COLUMNS = [
    ("vendor_code", "Артикул WB (vendorCode) — НЕ менять, это ключ"),
    ("title", "Название (до 60 символов — ограничение WB!)"),
    ("description", "Описание"),
    ("images", "Фото: ссылки через | , первая = главная (пусто = не менять)"),
    ("video_url", "Видео: ссылка на .mp4/.mov, до 50 МБ, максимум 1 (пусто = не менять)"),
    ("notes", "Заметки"),
]


def _cards_by_vendor(wb_data: dict) -> Dict[str, dict]:
    return {c.get("vendorCode"): c for c in wb_data.get("cards", []) if c.get("vendorCode")}


def _card_title(card: dict) -> str:
    return card.get("title") or card.get("subjectName") or ""


def _card_description(card: dict) -> str:
    return card.get("description") or ""


def _extract_photo_urls(card: dict) -> List[str]:
    """
    Собирает текущие ссылки на фото карточки. Поле в ответе WB может
    называться "photos" или "mediaFiles" в зависимости от версии ответа —
    проверьте на реальных данных после fetch-wb и поправьте здесь, если
    название поля другое.
    """
    urls: List[str] = []
    for key in ("photos", "mediaFiles", "media"):
        for item in card.get(key) or []:
            url = item.get("big") or item.get("c516x688") or item.get("url") if isinstance(item, dict) else item
            if url and url not in urls:
                urls.append(url)
        if urls:
            break
    return urls


def _extract_video_url(card: dict) -> str:
    """
    Пытается найти ссылку на уже загруженное видео карточки — НЕ ПРОВЕРЕНО
    на реальных данных (в живой документации WB не было явного примера, как
    видео выглядит в ответе /content/v2/get/cards/list; сюда собраны самые
    вероятные варианты названия поля). Если после fetch-wb колонка "Видео"
    у товаров с видео остаётся пустой — напишите мне, поправим по реальному
    JSON. ВАЖНО: пока это не проверено, если меняете фото у товара, у
    которого уже есть видео на WB, проверьте после push, что видео не
    пропало, и при необходимости впишите ссылку на него в колонку "Видео"
    вручную перед повторным push-wb-cards.
    """
    video = card.get("video") or card.get("videoUrl") or card.get("video_url")
    if isinstance(video, dict):
        return video.get("url") or video.get("link") or ""
    if isinstance(video, str):
        return video
    return ""


def build_wb_catalog(wb_data: dict, xlsx_path: str) -> int:
    """Строит редактируемый xlsx из уже загруженного data/wb_cards.json."""
    cards = _cards_by_vendor(wb_data)

    workbook = Workbook()
    ws = workbook.active
    ws.title = "WB"

    for col_idx, (_, header) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    row_idx = 2
    for vendor_code in sorted(cards.keys()):
        card = cards[vendor_code]
        title = _card_title(card)
        description = _card_description(card)
        images_str = "|".join(_extract_photo_urls(card))
        video_str = _extract_video_url(card)

        row_values = [vendor_code, title, description, images_str, video_str, ""]
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            if col_idx == 2 and len(title) > TITLE_MAX_LEN:
                cell.fill = WARN_FILL
        row_idx += 1

    widths = [18, 45, 45, 55, 45, 25]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(row_idx - 1, 1)}"

    workbook.save(xlsx_path)
    return row_idx - 2


def load_wb_catalog_edits(xlsx_path: str) -> Dict[str, dict]:
    workbook = load_workbook(xlsx_path, data_only=True)
    ws = workbook.active
    edits: Dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        vendor_code = str(row[0]).strip()
        raw = {COLUMNS[i][0]: (row[i] if i < len(row) else None) for i in range(len(COLUMNS))}
        images_raw = raw.get("images") or ""
        images = [u.strip() for u in str(images_raw).split("|") if u.strip()]
        title_val = raw.get("title")
        video_val = raw.get("video_url")
        edits[vendor_code] = {
            "title": title_val.strip() if isinstance(title_val, str) else title_val,
            "description": raw.get("description") or "",
            "images": images,
            "video_url": video_val.strip() if isinstance(video_val, str) else (video_val or ""),
            "notes": raw.get("notes"),
        }
    return edits


def _photo_index(fname: str, vendor_code: str) -> int:
    prefix = vendor_code + "_"
    if fname.startswith(prefix):
        m = re.match(r"(\d+)", fname[len(prefix):])
        if m:
            return int(m.group(1))
    return 0


def attach_wb_photos(xlsx_path: str, photos_dir: str, raw_base_url: str) -> Dict[str, List[str]]:
    """
    То же самое, что attach_local_photos в catalog_editor.py (Ozon), но
    ключ — vendorCode. Если артикулы у вас общие для Ozon и WB (как и
    задумано), фото из той же папки photos/ подойдут сразу для обеих
    площадок — переименовывать/дублировать файлы не нужно.
    """
    workbook = load_workbook(xlsx_path)
    ws = workbook.active

    vendor_col_idx = 1
    images_col_idx = next(i for i, (key, _) in enumerate(COLUMNS, start=1) if key == "images")

    files = [f for f in os.listdir(photos_dir) if os.path.splitext(f)[1].lower() in IMAGE_EXTS]

    matched: Dict[str, List[str]] = {}
    for row in ws.iter_rows(min_row=2):
        vendor_cell = row[vendor_col_idx - 1]
        if not vendor_cell.value:
            continue
        vendor_code = str(vendor_cell.value).strip()

        own_files = [
            f
            for f in files
            if f.startswith(vendor_code + "_") or f == vendor_code + os.path.splitext(f)[1]
        ]
        if not own_files:
            continue
        own_files.sort(key=lambda f: _photo_index(f, vendor_code))

        urls = [f"{raw_base_url.rstrip('/')}/{quote(f)}" for f in own_files]
        row[images_col_idx - 1].value = "|".join(urls)
        matched[vendor_code] = urls

    workbook.save(xlsx_path)
    return matched


def build_wb_update_items(wb_data: dict, edits: Dict[str, dict]) -> List[dict]:
    """
    Собирает items для /content/v2/cards/update — правит ТОЛЬКО title и
    description, всё остальное (brand, dimensions, characteristics, sizes,
    kizMarked) копирует из уже выгруженной карточки без изменений. Фото
    сюда не входят — для них build_wb_media_updates.
    """
    cards = _cards_by_vendor(wb_data)
    items: List[dict] = []
    for vendor_code, edit in edits.items():
        card = cards.get(vendor_code)
        if not card:
            continue
        title = (edit.get("title") or _card_title(card) or "").strip()
        description = edit.get("description") or _card_description(card)
        items.append(
            {
                "nmID": card.get("nmID"),
                "vendorCode": vendor_code,
                "kizMarked": card.get("kizMarked", False),
                "brand": card.get("brand", ""),
                "title": title[:TITLE_MAX_LEN],
                "description": description,
                "dimensions": card.get("dimensions") or {},
                "characteristics": card.get("characteristics") or [],
                "sizes": card.get("sizes") or [],
            }
        )
    return items


def build_wb_media_updates(wb_data: dict, edits: Dict[str, dict]) -> List[dict]:
    """
    Собирает список {"nm_id":..., "vendor_code":..., "urls": [...]} для
    вызова wb_client.update_media — для строк, где указаны фото И/ИЛИ видео
    в xlsx (обе колонки пустые = не трогаем медиа этого товара вообще).

    ВАЖНО: WB заменяет фото+видео целиком одним списком — если товар уже
    отправлялся с видео, а сейчас меняете только фото, видео нужно оставить
    в колонке "Видео" (не стирать), иначе оно пропадёт при этом запросе.
    Видео, если указано, всегда добавляется В КОНЕЦ списка (после фото).
    """
    cards = _cards_by_vendor(wb_data)
    out: List[dict] = []
    for vendor_code, edit in edits.items():
        images = list(edit.get("images") or [])
        video_url = (edit.get("video_url") or "").strip()
        if not images and not video_url:
            continue
        card = cards.get(vendor_code)
        if not card or not card.get("nmID"):
            continue
        urls = images + ([video_url] if video_url else [])
        out.append({"nm_id": card["nmID"], "vendor_code": vendor_code, "urls": urls})
    return out
