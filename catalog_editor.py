"""
Построение и разбор редактируемого Excel-каталога Ozon для полного
редактирования карточек: название, описание, цена и ФОТО.

Как это работает:
1. data/ozon_products.json — сырые данные, которые Ozon отдал по API
   (fetch-ozon). Это источник истины для штрихкода, категории, размеров,
   веса и всех технических характеристик товара.
2. build-ozon-catalog читает этот JSON и строит data/ozon_catalog.xlsx —
   таблицу, которую можно свободно редактировать в Excel/Google Таблицах.
3. Вы правите нужные строки/колонки и заливаете xlsx обратно в data/ в
   репозитории на GitHub (Add file -> Upload files, с заменой старого файла).
   Другого способа передать правки в GitHub Actions нет.
4. push-ozon-cards-dryrun / push-ozon-cards склеивают сырые данные из JSON
   с вашими правками из xlsx и отправляют результат в Ozon.

ФОТО: чтобы указать новую или другую фотографию, в колонку "Фото" впишите
прямую ссылку на картинку в интернете (заканчивается на .jpg/.jpeg/.png),
через символ | если фото несколько. Первая ссылка станет главным фото.
Ссылка должна быть ОБЩЕДОСТУПНОЙ — Ozon must скачать картинку сам по этой
ссылке, поэтому пароли/личные облака не подойдут. Если оставить колонку
"Фото" пустой — старые фото товара останутся как есть, ничего не сотрётся.

ВАЖНО (full-replace): методы Ozon /v3/product/import и
/v1/product/pictures/import считаются "перезаписывающими" — то, что вы не
укажете в запросе, Ozon может посчитать пустым и стереть. Поэтому здесь
переопределяются ТОЛЬКО поля из редактируемых колонок (название, описание,
цена, старая цена, фото); всё остальное (штрихкод, категория, тип, размеры,
вес, прочие характеристики, complex_attributes, цветное фото) копируется из
сырых данных Ozon без изменений — так собственные данные карточки не
портятся при правке одного поля.
"""
import logging
import os
import re
from typing import Dict, List, Optional
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger("marketplace-agent.catalog_editor")

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
EDIT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
NORMAL_FONT = Font(name=FONT_NAME)

# Порядок колонок важен — load_catalog_edits читает их по позиции, а не по
# тексту заголовка (текст заголовка можно менять, порядок — нет).
COLUMNS = [
    ("offer_id", "Артикул (offer_id) — НЕ менять, это ключ"),
    ("name", "Название товара"),
    ("description", "Описание"),
    ("price", "Цена, ₽"),
    ("old_price", "Цена до скидки, ₽ (0 или пусто — без скидки)"),
    ("images", "Фото: ссылки через | , первая = главная (пусто = не менять)"),
    ("video_url", "Видео: ссылка на .mp4/.mov, до 50 МБ (пусто = не менять, ЭКСПЕРИМЕНТАЛЬНО см. README)"),
    ("quantity_to_sell", "Кол-во к продаже (пока не используется push-ozon-cards)"),
    ("notes", "Заметки"),
]

# id атрибута внутри блока attributes, который зеркалит название товара —
# подтверждено на 65 реальных товарах 02.09.2026 (совпадает с полем name).
ATTR_ID_NAME_MIRROR = 4180
# id атрибута, где хранится полное описание товара.
ATTR_ID_DESCRIPTION = 4191

IMAGE_EXTS = {".jpg", ".jpeg", ".png"}


def _photo_index(fname: str, offer_id: str) -> int:
    """
    Номер фото — это число сразу после "<offer_id>_" в НАЧАЛЕ имени файла.
    После номера можно дописать любое читаемое описание — оно не мешает
    сортировке, так как не участвует в разборе:
        143210608_1_общий-вид.jpg   -> номер 1
        143210608_2_короб.jpg       -> номер 2
        143210608_10_маркировка.jpg -> номер 10 (а не "встанет перед 2")
    Файл без номера ("143210608.jpg") получает номер 0 и идёт первым.
    """
    prefix = offer_id + "_"
    if fname.startswith(prefix):
        m = re.match(r"(\d+)", fname[len(prefix):])
        if m:
            return int(m.group(1))
    return 0


def attach_local_photos(xlsx_path: str, photos_dir: str, raw_base_url: str = "") -> Dict[str, List[str]]:
    """
    Подставляет в колонку "Фото" уже собранного xlsx ссылки на файлы из
    папки photos_dir — без ручного копирования ссылок по одной.

    Группировка по товару — через ИМЯ ФАЙЛА: все фото одного товара должны
    начинаться с его артикула (offer_id) и номера, например для товара
    143210608:
        143210608_1_общий-вид.jpg
        143210608_2_короб.jpg
        143210608_3_маркировка.jpg
    (порядок — по номеру после "_": _1 будет главным фото). Текст после
    номера — любое читаемое описание для вас самих (необязательно, но
    удобно: так видно, что за фото, просто взглянув на список имён файлов,
    не открывая ссылку). Если фото одно — можно назвать файл просто
    "143210608.jpg" (без номера).

    Строки, для которых в photos_dir не нашлось ни одного подходящего файла,
    НЕ ТРОГАЮТСЯ — их текущее содержимое колонки "Фото" остаётся как было.

    Каждый найденный файл загружается как ассет GitHub Release (photo_host.py)
    и подставляется уже готовая ссылка на скачивание — раньше здесь строилась
    ссылка на raw_base_url (raw.githubusercontent.com), но эта раздача файлов
    оказалась ненадёжной на практике (см. photo_host.py); параметр оставлен
    только для обратной совместимости вызова и не используется.

    Возвращает offer_id -> список подставленных ссылок (для вывода в лог).
    """
    import photo_host

    wb = load_workbook(xlsx_path)
    ws = wb.active

    offer_id_col_idx = 1
    images_col_idx = next(i for i, (key, _) in enumerate(COLUMNS, start=1) if key == "images")

    files = [f for f in os.listdir(photos_dir) if os.path.splitext(f)[1].lower() in IMAGE_EXTS]

    matched: Dict[str, List[str]] = {}
    for row in ws.iter_rows(min_row=2):
        offer_id_cell = row[offer_id_col_idx - 1]
        if not offer_id_cell.value:
            continue
        offer_id = str(offer_id_cell.value).strip()

        own_files = [
            f
            for f in files
            if f.startswith(offer_id + "_") or f == offer_id + os.path.splitext(f)[1]
        ]
        if not own_files:
            continue
        own_files.sort(key=lambda f: _photo_index(f, offer_id))

        urls = []
        for f in own_files:
            try:
                url = photo_host.upload_file(os.path.join(photos_dir, f))
            except Exception as exc:
                logger.warning("%s: не удалось загрузить фото %s: %s", offer_id, f, exc)
                continue
            urls.append(url)
        if not urls:
            continue
        row[images_col_idx - 1].value = "|".join(urls)
        matched[offer_id] = urls

    wb.save(xlsx_path)
    return matched


def _find_attr_value(attributes: Optional[List[dict]], attr_id: int) -> str:
    for attr in attributes or []:
        if attr.get("id") == attr_id:
            values = attr.get("values") or []
            if values:
                return values[0].get("value", "") or ""
    return ""


def _with_overridden_attr(attributes: Optional[List[dict]], attr_id: int, new_value: str) -> List[dict]:
    """
    Возвращает НОВЫЙ список attributes с изменённым текстом значения
    атрибута attr_id (dictionary_value_id, если был, сохраняется — меняется
    только value). Если такого атрибута не было, добавляет новый. Остальные
    атрибуты копируются без изменений.
    """
    result: List[dict] = []
    found = False
    for attr in attributes or []:
        attr = dict(attr)
        if attr.get("id") == attr_id:
            old_values = attr.get("values") or [{}]
            new_first = dict(old_values[0]) if old_values else {}
            new_first["value"] = new_value
            attr["values"] = [new_first] + [dict(v) for v in old_values[1:]]
            found = True
        result.append(attr)
    if not found:
        result.append({"id": attr_id, "complex_id": 0, "values": [{"value": new_value}]})
    return result


def _extract_images(name_entry: dict) -> List[str]:
    """Собирает упорядоченный список ссылок на фото: сначала primary_image, затем images."""
    urls: List[str] = []
    primary = name_entry.get("primary_image")
    if isinstance(primary, list):
        primary = primary[0] if primary else None
    if primary:
        urls.append(primary)
    for img in name_entry.get("images") or []:
        url = img.get("file_name") if isinstance(img, dict) else img
        if url and url not in urls:
            urls.append(url)
    return urls


def _price_block(detail: dict) -> dict:
    block = detail.get("price")
    return block if isinstance(block, dict) else detail


def build_ozon_catalog(data: dict, xlsx_path: str) -> int:
    """
    Строит редактируемый xlsx из уже загруженного data/ozon_products.json
    (словарь с ключами 'list', 'details', 'names'). Возвращает количество
    товаров, записанных в файл.
    """
    names_by_offer = {n.get("offer_id"): n for n in data.get("names", []) if n.get("offer_id")}
    details_by_offer = {d.get("offer_id"): d for d in data.get("details", []) if d.get("offer_id")}

    wb = Workbook()
    ws = wb.active
    ws.title = "Ozon"

    for col_idx, (_, header) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    row_idx = 2
    for offer_id in sorted(names_by_offer.keys()):
        n = names_by_offer[offer_id]
        d = details_by_offer.get(offer_id, {})
        price_block = _price_block(d)

        name = n.get("name") or _find_attr_value(n.get("attributes"), ATTR_ID_NAME_MIRROR)
        description = _find_attr_value(n.get("attributes"), ATTR_ID_DESCRIPTION)
        price = price_block.get("price", "")
        old_price = price_block.get("old_price", "")
        images_str = "|".join(_extract_images(n))

        row_values = [offer_id, name, description, price, old_price, images_str, "", "", ""]
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            if COLUMNS[col_idx - 1][0] == "quantity_to_sell":
                cell.fill = EDIT_FILL
        row_idx += 1

    widths = [18, 45, 45, 12, 18, 55, 45, 20, 25]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(row_idx - 1, 1)}"

    wb.save(xlsx_path)
    return row_idx - 2


def load_catalog_edits(xlsx_path: str) -> Dict[str, dict]:
    """
    Читает отредактированный каталог обратно.
    Возвращает offer_id -> {"name", "description", "price", "old_price",
    "images": [...], "quantity_to_sell", "notes"}.
    Колонки читаются по позиции (см. COLUMNS), заголовки можно менять текстом.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    edits: Dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        offer_id = str(row[0]).strip()
        if not offer_id:
            continue
        raw = {COLUMNS[i][0]: (row[i] if i < len(row) else None) for i in range(len(COLUMNS))}
        images_raw = raw.get("images") or ""
        images = [u.strip() for u in str(images_raw).split("|") if u.strip()]
        name_val = raw.get("name")
        video_val = raw.get("video_url")
        edits[offer_id] = {
            "name": name_val.strip() if isinstance(name_val, str) else name_val,
            "description": raw.get("description") or "",
            "price": raw.get("price"),
            "old_price": raw.get("old_price"),
            "images": images,
            "video_url": video_val.strip() if isinstance(video_val, str) else (video_val or ""),
            "quantity_to_sell": raw.get("quantity_to_sell"),
            "notes": raw.get("notes"),
        }
    return edits


# --- Видео (ЭКСПЕРИМЕНТАЛЬНО, см. README) ---------------------------------
#
# У Ozon нет отдельного метода для видео, как у WB (/content/v3/media/save).
# Судя по документации, ссылка на видео передаётся как значение ОДНОЙ ИЗ
# ХАРАКТЕРИСТИК (attributes) конкретной категории товара — то есть у каждой
# категории может быть (или не быть) своя характеристика с названием вроде
# "Видео" / "Ссылка на видео". Ниже — попытка найти такую характеристику
# автоматически по названию через живой список характеристик категории
# (get_category_attributes). НЕ ПРОВЕРЕНО на реальных данных — при первом
# использовании смотрите вывод push-ozon-cards-dryrun: если для товара
# написано "атрибут 'видео' не найден", значит в этой категории Ozon видео
# через API добавить нельзя (или характеристика называется иначе — тогда
# напишите мне точное название из личного кабинета Ozon, поправим поиск).
_category_attr_cache: Dict[tuple, List[dict]] = {}


def _category_attrs(description_category_id: int, type_id: int) -> List[dict]:
    import ozon_client

    key = (description_category_id, type_id)
    if key not in _category_attr_cache:
        try:
            _category_attr_cache[key] = ozon_client.get_category_attributes(description_category_id, type_id)
        except Exception as exc:  # noqa: BLE001 — не хотим ронять весь push из-за одного запроса
            logger.warning(
                "Не удалось получить характеристики категории %s/%s: %s",
                description_category_id,
                type_id,
                exc,
            )
            _category_attr_cache[key] = []
    return _category_attr_cache[key]


def _find_category_attr_id(description_category_id: int, type_id: int, *keywords: str) -> Optional[int]:
    """Ищет id характеристики категории по вхождению одного из keywords в название (без учёта регистра)."""
    keywords_lower = [k.lower() for k in keywords]
    for attr in _category_attrs(description_category_id, type_id):
        name = (attr.get("name") or "").lower()
        if any(k in name for k in keywords_lower):
            return attr.get("id")
    return None


def _get_video_attr_id(description_category_id: int, type_id: int) -> Optional[int]:
    return _find_category_attr_id(description_category_id, type_id, "видео", "video")


def _get_tnved_attr_id(description_category_id: int, type_id: int) -> Optional[int]:
    return _find_category_attr_id(description_category_id, type_id, "тн вэд", "tn ved", "тнвэд")


def _build_one_item(name_entry: dict, detail: dict, edit: dict) -> dict:
    price_block = _price_block(detail)

    new_name = (edit.get("name") or name_entry.get("name") or "").strip()
    new_description = edit.get("description") or _find_attr_value(
        name_entry.get("attributes"), ATTR_ID_DESCRIPTION
    )

    attributes = name_entry.get("attributes") or []
    attributes = _with_overridden_attr(attributes, ATTR_ID_NAME_MIRROR, new_name)
    attributes = _with_overridden_attr(attributes, ATTR_ID_DESCRIPTION, new_description)

    video_url = (edit.get("video_url") or "").strip()
    if video_url:
        video_attr_id = _get_video_attr_id(
            name_entry.get("description_category_id", 0), name_entry.get("type_id", 0)
        )
        if video_attr_id:
            attributes = _with_overridden_attr(attributes, video_attr_id, video_url)
        else:
            logger.warning(
                "%s: в характеристиках его категории не нашлось атрибута 'видео' — видео "
                "НЕ добавлено, остальное обновится как обычно.",
                name_entry.get("offer_id"),
            )

    edited_images = edit.get("images") or []
    final_images = edited_images if edited_images else _extract_images(name_entry)
    primary_image = final_images[0] if final_images else ""
    images = final_images[1:] if len(final_images) > 1 else []

    price = edit.get("price")
    price = str(price).strip() if price not in (None, "") else str(price_block.get("price", "0"))
    old_price = edit.get("old_price")
    old_price = (
        str(old_price).strip() if old_price not in (None, "") else str(price_block.get("old_price", "0"))
    )
    vat = price_block.get("vat") or "0"
    currency_code = price_block.get("currency_code") or detail.get("currency_code") or "RUB"

    return {
        "offer_id": name_entry.get("offer_id", ""),
        "name": new_name,
        "description_category_id": name_entry.get("description_category_id", 0),
        "new_description_category_id": 0,
        "type_id": name_entry.get("type_id", 0),
        "barcode": name_entry.get("barcode", "") or "",
        "attributes": attributes,
        "complex_attributes": name_entry.get("complex_attributes") or [],
        "color_image": name_entry.get("color_image", "") or "",
        "images": images,
        "primary_image": primary_image,
        "pdf_list": name_entry.get("pdf_list") or [],
        "currency_code": currency_code,
        "price": price,
        "old_price": old_price,
        "vat": str(vat),
        "depth": name_entry.get("depth", 0),
        "width": name_entry.get("width", 0),
        "height": name_entry.get("height", 0),
        "dimension_unit": name_entry.get("dimension_unit") or "mm",
        "weight": name_entry.get("weight", 0),
        "weight_unit": name_entry.get("weight_unit") or "g",
    }


def build_import_items(data: dict, edits: Dict[str, dict]) -> List[dict]:
    """Склеивает сырые данные Ozon с правками из xlsx в готовые items для /v3/product/import."""
    names_by_offer = {n.get("offer_id"): n for n in data.get("names", []) if n.get("offer_id")}
    details_by_offer = {d.get("offer_id"): d for d in data.get("details", []) if d.get("offer_id")}

    items: List[dict] = []
    for offer_id, edit in edits.items():
        name_entry = names_by_offer.get(offer_id)
        if not name_entry:
            logger.warning(
                "Артикул %s есть в xlsx, но не найден в data/ozon_products.json — пропущен "
                "(похоже на опечатку в offer_id или устаревший fetch-ozon).",
                offer_id,
            )
            continue
        detail = details_by_offer.get(offer_id, {})
        items.append(_build_one_item(name_entry, detail, edit))
    return items
